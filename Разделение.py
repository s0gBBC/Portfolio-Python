#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import os
import pandas as pd
from tkinter import Tk, Listbox, Scrollbar, Button, Label, Frame, Checkbutton, IntVar, Canvas
from tkinter.filedialog import askopenfilename, askdirectory

class ExcelSplitterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Разбиение Excel файла")

        # Инициализация всех атрибутов
        self.file_path = None
        self.sheet_name = None
        self.column_name = None
        self.selected_values = []

        self.create_widgets()

    def create_widgets(self):
        """Создание всех виджетов на одном экране"""
        # Фрейм для выбора файла
        self.frame = Frame(self.root)
        self.frame.pack(padx=10, pady=10)

        # Кнопка для выбора файла
        self.file_button = Button(self.frame, text="Выбрать файл", command=self.select_file)
        self.file_button.grid(row=0, column=0, padx=5, pady=5)

        self.status_label = Label(self.frame, text="Выберите файл для обработки.")
        self.status_label.grid(row=1, column=0, padx=5, pady=5)

    def select_file(self):
        """Функция для выбора файла"""
        self.file_path = askopenfilename(title="Выберите файл", filetypes=[("Excel files", "*.xlsx")])
        if self.file_path:
            self.status_label.config(text="Выберите лист для обработки.")
            self.file_button.config(state="disabled")  # Отключаем кнопку выбора файла
            self.select_sheet()
        else:
            self.status_label.config(text="Файл не выбран.")

    def select_sheet(self):
        """Выбор листа из файла"""
        # Чтение данных из файла
        self.xls = pd.ExcelFile(self.file_path)
        self.sheet_names = self.xls.sheet_names  # Получаем список всех листов

        # Фрейм для отображения списка листов
        self.clear_frame()

        Label(self.frame, text="Выберите лист для обработки:").grid(row=0, column=0, padx=5, pady=5)

        # Список листов
        self.sheet_listbox = Listbox(self.frame, height=5, width=50)
        for sheet in self.sheet_names:
            self.sheet_listbox.insert("end", sheet)
        self.sheet_listbox.grid(row=1, column=0, padx=5, pady=5)

        # Добавление скроллбара
        scrollbar = Scrollbar(self.frame, orient="vertical", command=self.sheet_listbox.yview)
        scrollbar.grid(row=1, column=1, padx=5, pady=5, sticky="ns")
        self.sheet_listbox.config(yscrollcommand=scrollbar.set)

        # Кнопка для продолжения
        self.continue_button = Button(self.frame, text="Продолжить", command=self.select_column)
        self.continue_button.grid(row=2, column=0, padx=5, pady=5)

        # Кнопка назад
        self.back_button = Button(self.frame, text="Назад", command=self.back_to_file_selection)
        self.back_button.grid(row=2, column=1, padx=5, pady=5)

    def select_column(self):
        """Выбор столбца из выбранного листа"""
        selected_index = self.sheet_listbox.curselection()
        if not selected_index:
            self.status_label.config(text="Лист не выбран.")
            return

        self.sheet_name = self.sheet_listbox.get(selected_index)  # Получаем выбранный лист

        # Чтение данных для выбранного листа
        self.df = pd.read_excel(self.file_path, sheet_name=self.sheet_name)

        # Получаем список столбцов
        columns = self.df.columns.tolist()

        # Фрейм для отображения списка столбцов
        self.clear_frame()

        Label(self.frame, text="Выберите столбец для разделения:").grid(row=0, column=0, padx=5, pady=5)

        # Список столбцов
        self.column_listbox = Listbox(self.frame, height=10, width=50, selectmode="single")
        for column in columns:
            self.column_listbox.insert("end", column)
        self.column_listbox.grid(row=1, column=0, padx=5, pady=5)

        # Добавление скроллбара
        scrollbar = Scrollbar(self.frame, orient="vertical", command=self.column_listbox.yview)
        scrollbar.grid(row=1, column=1, padx=5, pady=5, sticky="ns")
        self.column_listbox.config(yscrollcommand=scrollbar.set)

        # Кнопка для продолжения
        self.continue_button = Button(self.frame, text="Продолжить", command=self.select_values)
        self.continue_button.grid(row=2, column=0, padx=5, pady=5)

        # Кнопка назад
        self.back_button = Button(self.frame, text="Назад", command=self.select_sheet)
        self.back_button.grid(row=2, column=1, padx=5, pady=5)

    def select_values(self):
        """Выбор значений из выбранного столбца"""
        selected_index = self.column_listbox.curselection()
        if not selected_index:
            self.status_label.config(text="Столбец не выбран.")
            return

        self.column_name = self.column_listbox.get(selected_index)  # Получаем выбранный столбец

        # Получаем уникальные значения в столбце
        unique_values = self.df[self.column_name].unique()

        # Фрейм для отображения списка значений
        self.clear_frame()

        Label(self.frame, text="Выберите значения для фильтрации:").grid(row=0, column=0, padx=5, pady=5)

        # Создание canvas и scroll для прокрутки
        canvas = Canvas(self.frame)
        canvas.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")

        scrollbar = Scrollbar(self.frame, orient="vertical", command=canvas.yview)
        scrollbar.grid(row=1, column=1, padx=5, pady=5, sticky="ns")

        canvas.config(yscrollcommand=scrollbar.set)

        # Фрейм внутри canvas для размещения чекбоксов
        check_frame = Frame(canvas)
        canvas.create_window((0, 0), window=check_frame, anchor="nw")

        # Список чекбоксов для значений
        self.var_dict = {}
        self.checkbuttons = []
        for value in unique_values:
            var = IntVar(value=1)  # Значение по умолчанию - отмечено
            checkbutton = Checkbutton(check_frame, text=str(value), variable=var)
            checkbutton.pack(anchor="w")
            self.checkbuttons.append(checkbutton)
            self.var_dict[value] = var

        # Чекбокс "Выбрать все"
        self.select_all_var = IntVar()
        select_all_checkbox = Checkbutton(self.frame, text="Выбрать все", variable=self.select_all_var, command=self.select_all)
        select_all_checkbox.grid(row=2, column=0, padx=5, pady=5)

        # Кнопка для продолжения
        self.continue_button = Button(self.frame, text="Продолжить", command=self.split_file)
        self.continue_button.grid(row=3, column=0, padx=5, pady=5)

        # Кнопка назад
        self.back_button = Button(self.frame, text="Назад", command=self.select_column)
        self.back_button.grid(row=3, column=1, padx=5, pady=5)

        # Обновляем размер canvas для правильного отображения прокрутки
        check_frame.update_idletasks()
        canvas.config(scrollregion=canvas.bbox("all"))

    def select_all(self):
        """Функция для отметки всех чекбоксов"""
        select_all = self.select_all_var.get()
        for var in self.var_dict.values():
            var.set(select_all)  # Устанавливаем одинаковое значение для всех чекбоксов

    def split_file(self):
        """Разбиение файла по выбранным значениям"""
        selected_values = [value for value, var in self.var_dict.items() if var.get() == 1]
        if not selected_values:
            self.status_label.config(text="Значения не выбраны.")
            return

        # Создание папки для сохранения
        output_dir = askdirectory(title="Выберите папку для сохранения файлов")
        if not output_dir:
            self.status_label.config(text="Папка не выбрана. Программа завершена.")
            return

        # Фильтруем данные по выбранным значениям
        filtered_df = self.df[self.df[self.column_name].isin(selected_values)]

        if filtered_df.empty:
            self.status_label.config(text="Нет данных для выбранных значений.")
            return

        # Сохраняем разделённый файл
        for value in selected_values:
            value_df = filtered_df[filtered_df[self.column_name] == value]
            report_path = os.path.join(output_dir, f"{self.sheet_name}_{value}.xlsx")
            value_df.to_excel(report_path, index=False)

        self.status_label.config(text="Процесс завершен!")
        self.root.after(2000, self.root.quit)  # Закрытие программы через 2 секунды

    def clear_frame(self):
        """Очистка фрейма перед добавлением новых элементов"""
        for widget in self.frame.winfo_children():
            widget.destroy()


# Запуск программы
root = Tk()
app = ExcelSplitterApp(root)
root.mainloop()


# # Свод всех файлов в один

# In[2]:


import os
import pandas as pd
from tkinter import Tk, Listbox, Scrollbar, Button, Label, Frame, filedialog, messagebox
from tkinter.filedialog import askopenfilenames, askdirectory

class ExcelMergerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Объединение Excel файлов")

        # Инициализация атрибутов
        self.file_paths = []
        self.sheet_name = None
        self.key_columns = []
        self.merge_columns = []

        self.create_widgets()

    def create_widgets(self):
        """Создание всех виджетов на одном экране"""
        self.frame = Frame(self.root)
        self.frame.pack(padx=10, pady=10)

        # Кнопка для выбора файлов
        self.files_button = Button(self.frame, text="Выбрать файлы", command=self.select_files)
        self.files_button.grid(row=0, column=0, padx=5, pady=5)

        self.status_label = Label(self.frame, text="Выберите файлы для объединения.")
        self.status_label.grid(row=1, column=0, padx=5, pady=5)

    def select_files(self):
        """Функция для выбора файлов"""
        self.file_paths = askopenfilenames(
            title="Выберите файлы", 
            filetypes=[("Excel files", "*.xlsx;*.xls")]
        )
        
        if not self.file_paths:
            self.status_label.config(text="Файлы не выбраны.")
            return
            
        self.status_label.config(text=f"Выбрано {len(self.file_paths)} файлов.")
        self.files_button.config(state="disabled")
        
        # Проверяем, есть ли общий лист во всех файлах
        common_sheets = self.find_common_sheets()
        
        if not common_sheets:
            messagebox.showerror("Ошибка", "Нет общих листов во всех выбранных файлах.")
            return
            
        self.select_sheet(common_sheets)

    def find_common_sheets(self):
        """Находит общие листы во всех выбранных файлах"""
        sheets_list = []
        for file_path in self.file_paths:
            xls = pd.ExcelFile(file_path)
            sheets_list.append(set(xls.sheet_names))
            
        common_sheets = set.intersection(*sheets_list)
        return list(common_sheets)

    def select_sheet(self, sheet_names):
        """Выбор листа из файла"""
        self.clear_frame()

        Label(self.frame, text="Выберите лист для объединения:").grid(row=0, column=0, padx=5, pady=5)

        # Список листов
        self.sheet_listbox = Listbox(self.frame, height=5, width=50)
        for sheet in sheet_names:
            self.sheet_listbox.insert("end", sheet)
        self.sheet_listbox.grid(row=1, column=0, padx=5, pady=5)

        # Скроллбар
        scrollbar = Scrollbar(self.frame, orient="vertical", command=self.sheet_listbox.yview)
        scrollbar.grid(row=1, column=1, padx=5, pady=5, sticky="ns")
        self.sheet_listbox.config(yscrollcommand=scrollbar.set)

        # Кнопка для продолжения
        Button(self.frame, text="Продолжить", command=self.prepare_merge).grid(row=2, column=0, padx=5, pady=5)
        Button(self.frame, text="Назад", command=self.create_widgets).grid(row=2, column=1, padx=5, pady=5)

    def prepare_merge(self):
        """Подготовка к объединению - выбор ключевых столбцов"""
        selected_index = self.sheet_listbox.curselection()
        if not selected_index:
            self.status_label.config(text="Лист не выбран.")
            return

        self.sheet_name = self.sheet_listbox.get(selected_index)
        
        # Загружаем первый файл для получения столбцов
        df = pd.read_excel(self.file_paths[0], sheet_name=self.sheet_name)
        columns = df.columns.tolist()

        self.clear_frame()

        Label(self.frame, text="Выберите ключевые столбцы для объединения:").grid(row=0, column=0, padx=5, pady=5)
        Label(self.frame, text="(По этим столбцам будет происходить сопоставление строк)").grid(row=1, column=0, padx=5, pady=5)

        # Список столбцов (множественный выбор)
        self.columns_listbox = Listbox(self.frame, height=10, width=50, selectmode="multiple")
        for column in columns:
            self.columns_listbox.insert("end", column)
        self.columns_listbox.grid(row=2, column=0, padx=5, pady=5)

        # Скроллбар
        scrollbar = Scrollbar(self.frame, orient="vertical", command=self.columns_listbox.yview)
        scrollbar.grid(row=2, column=1, padx=5, pady=5, sticky="ns")
        self.columns_listbox.config(yscrollcommand=scrollbar.set)

        # Кнопки
        Button(self.frame, text="Продолжить", command=self.process_merge).grid(row=3, column=0, padx=5, pady=5)
        Button(self.frame, text="Назад", command=lambda: self.select_sheet(self.find_common_sheets())).grid(row=3, column=1, padx=5, pady=5)

    def process_merge(self):
        """Процесс объединения файлов"""
        selected_indices = self.columns_listbox.curselection()
        if not selected_indices:
            messagebox.showerror("Ошибка", "Не выбраны ключевые столбцы.")
            return
            
        self.key_columns = [self.columns_listbox.get(i) for i in selected_indices]
        
        # Загружаем все файлы
        dfs = []
        for file_path in self.file_paths:
            df = pd.read_excel(file_path, sheet_name=self.sheet_name)
            dfs.append(df)
            
        # Объединяем данные
        merged_df = self.merge_dataframes(dfs)
        
        # Сохраняем результат
        output_dir = askdirectory(title="Выберите папку для сохранения результата")
        if not output_dir:
            self.status_label.config(text="Папка не выбрана.")
            return
            
        output_path = os.path.join(output_dir, f"merged_{self.sheet_name}.xlsx")
        merged_df.to_excel(output_path, index=False)
        
        messagebox.showinfo("Успех", f"Файлы успешно объединены и сохранены как:\n{output_path}")
        self.root.after(2000, self.root.quit)

    def merge_dataframes(self, dfs):
        """
        Объединяет DataFrame, заполняя недостающие данные из других DataFrame
        """
        # Начинаем с первого DataFrame
        merged_df = dfs[0].copy()
        
        # Для каждого последующего DataFrame
        for df in dfs[1:]:
            # Объединяем по ключевым столбцам
            for _, row in df.iterrows():
                # Ищем совпадение в merged_df по ключевым столбцам
                mask = pd.Series(True, index=merged_df.index)
                for col in self.key_columns:
                    mask &= (merged_df[col] == row[col])
                
                if mask.any():  # Если нашли совпадение
                    idx = mask.idxmax()
                    # Обновляем только пустые или NaN значения
                    for col in df.columns:
                        if col not in self.key_columns:
                            if pd.isna(merged_df.at[idx, col]) or merged_df.at[idx, col] == '':
                                merged_df.at[idx, col] = row[col]
                else:  # Если не нашли - добавляем новую строку
                    merged_df = pd.concat([merged_df, row.to_frame().T], ignore_index=True)
        
        return merged_df

    def clear_frame(self):
        """Очистка фрейма перед добавлением новых элементов"""
        for widget in self.frame.winfo_children():
            widget.destroy()


# Запуск программы
if __name__ == "__main__":
    root = Tk()
    app = ExcelMergerApp(root)
    root.mainloop()


# # Обновление данных в общем файле

# In[6]:


import os
import pandas as pd
from tkinter import Tk, Listbox, Scrollbar, Button, Label, Frame, filedialog, messagebox, ttk
from tkinter.filedialog import askopenfilename, askopenfilenames, asksaveasfilename
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

class ExcelUpdaterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Обновление Excel файла с подсветкой изменений")

        # Инициализация атрибутов
        self.base_file_path = None
        self.update_files_paths = []
        self.sheet_name = None
        self.key_columns = []
        self.changes_count = 0
        
        self.create_widgets()

    def create_widgets(self):
        """Создание всех виджетов"""
        self.frame = Frame(self.root)
        self.frame.pack(padx=10, pady=10)

        # Выбор базового файла
        self.base_file_button = Button(self.frame, text="1. Выбрать файл для обновления", 
                                     command=self.select_base_file)
        self.base_file_button.pack(pady=5)
        
        self.base_file_label = Label(self.frame, text="Файл не выбран", fg="gray")
        self.base_file_label.pack(pady=5)

        # Выбор файлов с обновлениями
        self.update_files_button = Button(self.frame, text="2. Выбрать файлы с обновлениями", 
                                        command=self.select_update_files, state="disabled")
        self.update_files_button.pack(pady=5)
        
        self.update_files_label = Label(self.frame, text="Файлы не выбраны", fg="gray")
        self.update_files_label.pack(pady=5)

        # Кнопка выбора ключевых столбцов
        self.select_columns_button = Button(self.frame, text="3. Выбрать ключевые столбцы", 
                                          command=self.select_key_columns, state="disabled")
        self.select_columns_button.pack(pady=5)

        # Кнопка запуска процесса
        self.run_button = Button(self.frame, text="4. Обновить данные с подсветкой", 
                                command=self.run_update, state="disabled")
        self.run_button.pack(pady=20)

        # Статус
        self.status_label = Label(self.frame, text="Выберите файлы для работы")
        self.status_label.pack(pady=10)

    def select_base_file(self):
        """Выбор базового файла"""
        self.base_file_path = askopenfilename(
            title="Выберите файл для обновления", 
            filetypes=[("Excel files", "*.xlsx;*.xls")]
        )
        
        if self.base_file_path:
            self.base_file_label.config(
                text=os.path.basename(self.base_file_path),
                fg="green"
            )
            self.update_files_button.config(state="normal")
            self.status_label.config(text="Теперь выберите файлы с обновлениями")
        else:
            self.base_file_label.config(text="Файл не выбран", fg="red")

    def select_update_files(self):
        """Выбор файлов с обновлениями"""
        self.update_files_paths = askopenfilenames(
            title="Выберите файлы с обновлениями", 
            filetypes=[("Excel files", "*.xlsx;*.xls")]
        )
        
        if self.update_files_paths:
            self.update_files_label.config(
                text=f"Выбрано {len(self.update_files_paths)} файлов",
                fg="green"
            )
            self.select_columns_button.config(state="normal")
            self.status_label.config(text="Теперь выберите ключевые столбцы")
        else:
            self.update_files_label.config(text="Файлы не выбраны", fg="red")

    def select_key_columns(self):
        """Выбор ключевых столбцов"""
        if not self.base_file_path:
            messagebox.showerror("Ошибка", "Сначала выберите базовый файл")
            return

        try:
            base_df = pd.read_excel(self.base_file_path)
            columns = base_df.columns.tolist()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить базовый файл:\n{e}")
            return

        # Создаем окно для выбора столбцов
        self.columns_window = Tk()
        self.columns_window.title("Выберите ключевые столбцы")
        
        Label(self.columns_window, text="Выберите столбцы для сопоставления строк:").pack(pady=5)
        Label(self.columns_window, text="(Обычно это регион и дата)").pack(pady=5)

        # Список столбцов
        self.columns_listbox = Listbox(self.columns_window, selectmode="multiple", height=10, width=40)
        for col in columns:
            self.columns_listbox.insert("end", col)
        self.columns_listbox.pack(pady=5, padx=10)

        # Кнопка подтверждения
        Button(self.columns_window, text="Подтвердить выбор", 
              command=self.confirm_columns).pack(pady=10)

    def confirm_columns(self):
        """Подтверждение выбранных столбцов"""
        selected_indices = self.columns_listbox.curselection()
        if not selected_indices:
            messagebox.showerror("Ошибка", "Не выбрано ни одного ключевого столбца")
            return
            
        self.key_columns = [self.columns_listbox.get(i) for i in selected_indices]
        self.columns_window.destroy()
        self.run_button.config(state="normal")
        self.status_label.config(text="Готово к обновлению. Нажмите 'Обновить данные с подсветкой'")

    def run_update(self):
        """Запуск процесса обновления с подсветкой изменений"""
        if not self.base_file_path or not self.update_files_paths or not self.key_columns:
            messagebox.showerror("Ошибка", "Сначала выполните все предыдущие шаги")
            return

        # Загружаем базовый файл
        try:
            base_df = pd.read_excel(self.base_file_path)
            original_base_df = base_df.copy()  # Сохраняем оригинальные данные для сравнения
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить базовый файл:\n{e}")
            return

        # Загружаем файлы с обновлениями
        update_dfs = []
        for file_path in self.update_files_paths:
            try:
                df = pd.read_excel(file_path)
                update_dfs.append(df)
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось загрузить файл {os.path.basename(file_path)}:\n{e}")
                return

        # Обновляем данные и получаем информацию об изменениях
        try:
            updated_df, change_log = self.update_dataframe_with_changes(base_df, update_dfs, original_base_df)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при обновлении данных:\n{e}")
            return

        # Предлагаем сохранить результат
        save_path = asksaveasfilename(
            title="Сохранить обновленный файл с подсветкой",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"updated_with_highlights_{os.path.basename(self.base_file_path)}"
        )
        
        if save_path:
            try:
                self.save_with_highlights(updated_df, change_log, save_path)
                messagebox.showinfo("Успех", 
                    f"Файл успешно обновлен с подсветкой изменений ({self.changes_count} изменений)\n"
                    f"Сохранен как:\n{save_path}")
                self.status_label.config(
                    text=f"Файл сохранен: {os.path.basename(save_path)} ({self.changes_count} изменений)",
                    fg="green"
                )
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось сохранить файл:\n{e}")
        else:
            self.status_label.config(text="Сохранение отменено", fg="orange")

    def update_dataframe_with_changes(self, base_df, update_dfs, original_df):
        """
        Обновляет DataFrame и возвращает лог изменений
        """
        updated_df = base_df.copy()
        change_log = []
        self.changes_count = 0
        
        # Для каждого DataFrame с обновлениями
        for df in update_dfs:
            # Проверяем наличие ключевых столбцов в файле обновления
            missing_key_cols = [col for col in self.key_columns if col not in df.columns]
            if missing_key_cols:
                messagebox.showwarning(
                    "Предупреждение", 
                    f"В файле {os.path.basename(df.attrs.get('file_path', ''))} отсутствуют ключевые столбцы: {', '.join(missing_key_cols)}\n"
                    "Эти строки не будут обновлены."
                )
                continue
            
            # Обновляем данные по ключевым столбцам
            for idx, row in df.iterrows():
                # Ищем совпадение в base_df по ключевым столбцам
                mask = pd.Series(True, index=updated_df.index)
                for col in self.key_columns:
                    mask &= (updated_df[col].astype(str) == str(row[col]))
                
                if mask.any():  # Если нашли совпадение
                    base_idx = mask.idxmax()
                    # Обновляем только непустые значения из файла обновления
                    for col in df.columns:
                        if col in updated_df.columns and col not in self.key_columns:
                            if not pd.isna(row[col]) and row[col] != '':
                                # Для строковых значений проверяем, что они не пустые
                                if isinstance(row[col], str) and not row[col].strip():
                                    continue
                                
                                # Проверяем, было ли изменение
                                original_value = original_df.at[base_idx, col]
                                new_value = row[col]
                                
                                if str(original_value) != str(new_value):
                                    # Записываем изменение
                                    change_log.append({
                                        'row': base_idx,
                                        'column': col,
                                        'old_value': original_value,
                                        'new_value': new_value
                                    })
                                    self.changes_count += 1
                                
                                updated_df.at[base_idx, col] = new_value
        
        return updated_df, change_log

    def save_with_highlights(self, df, change_log, save_path):
        """
        Сохраняет DataFrame с подсветкой измененных ячеек
        """
        # Создаем новую книгу Excel
        wb = Workbook()
        ws = wb.active
        
        # Заполняем лист данными из DataFrame
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Стиль для подсветки изменений
        highlight_fill = PatternFill(start_color="FFFF00",  # Желтый
                                    end_color="FFFF00",
                                    fill_type="solid")
        
        # Применяем подсветку к измененным ячейкам
        for change in change_log:
            # Находим столбец по имени
            col_letter = None
            for cell in ws[1]:  # Ищем в заголовках
                if cell.value == change['column']:
                    col_letter = cell.column_letter
                    break
            
            if col_letter:
                cell_ref = f"{col_letter}{change['row'] + 2}"  # +1 для заголовка, +1 для 1-based индексации
                ws[cell_ref].fill = highlight_fill
        
        # Сохраняем файл
        wb.save(save_path)

# Запуск программы
if __name__ == "__main__":
    root = Tk()
    app = ExcelUpdaterApp(root)
    root.mainloop()


# In[ ]:




