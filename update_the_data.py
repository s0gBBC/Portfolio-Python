import os
import pandas as pd
from tkinter import Tk, Listbox, Scrollbar, Button, Label, Frame, filedialog, messagebox, ttk
from tkinter.filedialog import askopenfilename, askopenfilenames, asksaveasfilename
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

class ExcelUpdaterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Обновление Excel файла с подсветкой изменений")

        # Инициализация атрибутов
        self.base_file_path = None
        self.update_files_paths = []
        self.sheet_name = None
        self.key_columns = []
        self.changes_count = 0
        
        self.create_widgets()

    def create_widgets(self):
        """Создание всех виджетов"""
        self.frame = Frame(self.root)
        self.frame.pack(padx=10, pady=10)

        # Выбор базового файла
        self.base_file_button = Button(self.frame, text="1. Выбрать файл для обновления", 
                                     command=self.select_base_file)
        self.base_file_button.pack(pady=5)
        
        self.base_file_label = Label(self.frame, text="Файл не выбран", fg="gray")
        self.base_file_label.pack(pady=5)

        # Выбор файлов с обновлениями
        self.update_files_button = Button(self.frame, text="2. Выбрать файлы с обновлениями", 
                                        command=self.select_update_files, state="disabled")
        self.update_files_button.pack(pady=5)
        
        self.update_files_label = Label(self.frame, text="Файлы не выбраны", fg="gray")
        self.update_files_label.pack(pady=5)

        # Кнопка выбора ключевых столбцов
        self.select_columns_button = Button(self.frame, text="3. Выбрать ключевые столбцы", 
                                          command=self.select_key_columns, state="disabled")
        self.select_columns_button.pack(pady=5)

        # Кнопка запуска процесса
        self.run_button = Button(self.frame, text="4. Обновить данные с подсветкой", 
                                command=self.run_update, state="disabled")
        self.run_button.pack(pady=20)

        # Статус
        self.status_label = Label(self.frame, text="Выберите файлы для работы")
        self.status_label.pack(pady=10)

    def select_base_file(self):
        """Выбор базового файла"""
        self.base_file_path = askopenfilename(
            title="Выберите файл для обновления", 
            filetypes=[("Excel files", "*.xlsx;*.xls")]
        )
        
        if self.base_file_path:
            self.base_file_label.config(
                text=os.path.basename(self.base_file_path),
                fg="green"
            )
            self.update_files_button.config(state="normal")
            self.status_label.config(text="Теперь выберите файлы с обновлениями")
        else:
            self.base_file_label.config(text="Файл не выбран", fg="red")

    def select_update_files(self):
        """Выбор файлов с обновлениями"""
        self.update_files_paths = askopenfilenames(
            title="Выберите файлы с обновлениями", 
            filetypes=[("Excel files", "*.xlsx;*.xls")]
        )
        
        if self.update_files_paths:
            self.update_files_label.config(
                text=f"Выбрано {len(self.update_files_paths)} файлов",
                fg="green"
            )
            self.select_columns_button.config(state="normal")
            self.status_label.config(text="Теперь выберите ключевые столбцы")
        else:
            self.update_files_label.config(text="Файлы не выбраны", fg="red")

    def select_key_columns(self):
        """Выбор ключевых столбцов"""
        if not self.base_file_path:
            messagebox.showerror("Ошибка", "Сначала выберите базовый файл")
            return

        try:
            base_df = pd.read_excel(self.base_file_path)
            columns = base_df.columns.tolist()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить базовый файл:\n{e}")
            return

        # Создаем окно для выбора столбцов
        self.columns_window = Tk()
        self.columns_window.title("Выберите ключевые столбцы")
        
        Label(self.columns_window, text="Выберите столбцы для сопоставления строк:").pack(pady=5)
        Label(self.columns_window, text="(Обычно это регион и дата)").pack(pady=5)

        # Список столбцов
        self.columns_listbox = Listbox(self.columns_window, selectmode="multiple", height=10, width=40)
        for col in columns:
            self.columns_listbox.insert("end", col)
        self.columns_listbox.pack(pady=5, padx=10)

        # Кнопка подтверждения
        Button(self.columns_window, text="Подтвердить выбор", 
              command=self.confirm_columns).pack(pady=10)

    def confirm_columns(self):
        """Подтверждение выбранных столбцов"""
        selected_indices = self.columns_listbox.curselection()
        if not selected_indices:
            messagebox.showerror("Ошибка", "Не выбрано ни одного ключевого столбца")
            return
            
        self.key_columns = [self.columns_listbox.get(i) for i in selected_indices]
        self.columns_window.destroy()
        self.run_button.config(state="normal")
        self.status_label.config(text="Готово к обновлению. Нажмите 'Обновить данные с подсветкой'")

    def run_update(self):
        """Запуск процесса обновления с подсветкой изменений"""
        if not self.base_file_path or not self.update_files_paths or not self.key_columns:
            messagebox.showerror("Ошибка", "Сначала выполните все предыдущие шаги")
            return

        # Загружаем базовый файл
        try:
            base_df = pd.read_excel(self.base_file_path)
            original_base_df = base_df.copy()  # Сохраняем оригинальные данные для сравнения
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить базовый файл:\n{e}")
            return

        # Загружаем файлы с обновлениями
        update_dfs = []
        for file_path in self.update_files_paths:
            try:
                df = pd.read_excel(file_path)
                update_dfs.append(df)
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось загрузить файл {os.path.basename(file_path)}:\n{e}")
                return

        # Обновляем данные и получаем информацию об изменениях
        try:
            updated_df, change_log = self.update_dataframe_with_changes(base_df, update_dfs, original_base_df)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при обновлении данных:\n{e}")
            return

        # Предлагаем сохранить результат
        save_path = asksaveasfilename(
            title="Сохранить обновленный файл с подсветкой",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"updated_with_highlights_{os.path.basename(self.base_file_path)}"
        )
        
        if save_path:
            try:
                self.save_with_highlights(updated_df, change_log, save_path)
                messagebox.showinfo("Успех", 
                    f"Файл успешно обновлен с подсветкой изменений ({self.changes_count} изменений)\n"
                    f"Сохранен как:\n{save_path}")
                self.status_label.config(
                    text=f"Файл сохранен: {os.path.basename(save_path)} ({self.changes_count} изменений)",
                    fg="green"
                )
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось сохранить файл:\n{e}")
        else:
            self.status_label.config(text="Сохранение отменено", fg="orange")

    def update_dataframe_with_changes(self, base_df, update_dfs, original_df):
        """
        Обновляет DataFrame и возвращает лог изменений
        """
        updated_df = base_df.copy()
        change_log = []
        self.changes_count = 0
        
        # Для каждого DataFrame с обновлениями
        for update_df in update_dfs:
            # Проверяем наличие ключевых столбцов в файле обновления
            missing_key_cols = [col for col in self.key_columns if col not in update_df.columns]
            if missing_key_cols:
                messagebox.showwarning(
                    "Предупреждение", 
                    f"В файле обновления отсутствуют ключевые столбцы: {', '.join(missing_key_cols)}\n"
                    "Эти строки не будут обновлены."
                )
                continue
            
            # Обновляем данные по ключевым столбцам
            for idx, update_row in update_df.iterrows():
                # Ищем совпадение в base_df по ключевым столбцам
                mask = pd.Series(True, index=updated_df.index)
                for col in self.key_columns:
                    mask &= (updated_df[col].astype(str) == str(update_row[col]))
                
                if mask.any():  # Если нашли совпадение
                    base_idx = mask.idxmax()
                    
                    # Обновляем только непустые значения из файла обновления
                    for col in update_df.columns:
                        if col in updated_df.columns and col not in self.key_columns:
                            # Пропускаем пустые значения
                            if pd.isna(update_row[col]) or update_row[col] == '':
                                continue
                            
                            # Для строковых значений проверяем, что они не пустые
                            if isinstance(update_row[col], str) and not update_row[col].strip():
                                continue
                            
                            # Получаем ОРИГИНАЛЬНОЕ значение до любых изменений
                            original_value = original_df.at[base_idx, col]
                            new_value = update_row[col]
                            
                            # Сравниваем как строки для надежности
                            if str(original_value).strip() != str(new_value).strip():
                                # Записываем изменение ДО обновления
                                change_log.append({
                                    'row': base_idx,
                                    'column': col,
                                    'old_value': original_value,
                                    'new_value': new_value
                                })
                                self.changes_count += 1
                                
                                # Обновляем значение в базовом DataFrame
                                updated_df.at[base_idx, col] = new_value
        
        return updated_df, change_log

    def save_with_highlights(self, df, change_log, save_path):
        """
        Сохраняет DataFrame с подсветкой измененных ячеек
        """
        # Создаем новую книгу Excel
        wb = Workbook()
        ws = wb.active
        
        # Заполняем лист данными из DataFrame
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Стиль для подсветки изменений
        highlight_fill = PatternFill(start_color="FFFF00",  # Желтый
                                    end_color="FFFF00",
                                    fill_type="solid")
        
        # Применяем подсветку к измененным ячейкам
        for change in change_log:
            # Находим столбец по имени
            col_letter = None
            for col_idx, cell in enumerate(ws[1], 1):  # Ищем в заголовках
                if cell.value == change['column']:
                    col_letter = cell.column_letter
                    break
            
            if col_letter:
                cell_ref = f"{col_letter}{change['row'] + 2}"  # +1 для заголовка, +1 для 1-based индексации
                # Проверяем, что значение в ячейке соответствует новому значению из лога
                if str(ws[cell_ref].value).strip() == str(change['new_value']).strip():
                    ws[cell_ref].fill = highlight_fill
        
        # Сохраняем файл
        wb.save(save_path)

# Запуск программы
if __name__ == "__main__":
    root = Tk()
    app = ExcelUpdaterApp(root)
    root.mainloop()